#!/usr/bin/env python3
"""
Excel 자산관리시트 데이터를 완전한 아카이브 시스템으로 변환하는 통합 스크립트
- 자산 데이터 (현금, 투자, 연금): 새로운 아카이브 시스템 (monthly_archives + assets_archive_data)
- 지출 데이터 (고정지출, 선납지출): 기존 아카이브 시스템 (monthly_archives + *_expenses_archive)
- 최신 월: 실시간 테이블에 입력
- 과거 월들: 아카이브 시스템에 저장
"""

import openpyxl
import mysql.connector
import re
import json
from datetime import datetime
import os

class ExcelDataConverter:
    def __init__(self):
        self.excel_file = '자산관리시트.xlsx'
        self.connection = None
        # 현재 월 설정 (가장 최신 데이터로 사용할 월)
        self.current_month = '202509'  # 필요시 수정

    def connect_db(self):
        """데이터베이스 연결"""
        try:
            self.connection = mysql.connector.connect(
                host='localhost',
                port=3306,
                user='root',
                password='money_dev_2024',
                database='money_management',
                charset='utf8mb4'
            )
            print("✅ 데이터베이스 연결 성공")
            return True
        except Exception as e:
            print(f"❌ 데이터베이스 연결 실패: {e}")
            return False

    def parse_excel_data(self):
        """Excel 파일에서 모든 월 데이터 파싱 (자산 + 지출)"""
        if not os.path.exists(self.excel_file):
            print(f"❌ Excel 파일을 찾을 수 없습니다: {self.excel_file}")
            return {}

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        all_data = {}

        print("📊 Excel 데이터 파싱 중...")

        for sheet_name in workbook.sheetnames:
            if len(sheet_name) == 6 and sheet_name.isdigit():  # YYYYMM 형식
                year = int(sheet_name[:4])
                month = int(sheet_name[4:])
                print(f"   📅 {year}년 {month}월 파싱 중...")

                sheet_data = self.parse_sheet_data(workbook[sheet_name])
                all_data[sheet_name] = {
                    'year': year,
                    'month': month,
                    'assets': sheet_data['assets'],
                    'expenses': sheet_data['expenses']
                }

                # 파싱 결과 출력
                asset_counts = {k: len(v) for k, v in sheet_data['assets'].items()}
                expense_counts = {k: len(v) for k, v in sheet_data['expenses'].items()}
                print(f"      ✅ 자산: 현금{asset_counts['cash_assets']}개, 투자{asset_counts['investment_assets']}개, 연금{asset_counts['pension_assets']}개")
                print(f"      ✅ 지출: 고정{expense_counts['fixed_expenses']}개, 선납{expense_counts['prepaid_expenses']}개")

        print(f"✅ 총 {len(all_data)}개월 데이터 파싱 완료")
        return all_data

    def parse_sheet_data(self, sheet):
        """개별 시트 데이터 파싱 (자산 + 지출)"""
        data = {
            'assets': {
                'cash_assets': [],
                'investment_assets': [],
                'pension_assets': []
            },
            'expenses': {
                'fixed_expenses': [],
                'prepaid_expenses': []
            }
        }

        # 자산 데이터 파싱 (A~H열)
        self.parse_assets_data(sheet, data['assets'])

        # 지출 데이터 파싱 (I~M열)
        self.parse_expenses_data(sheet, data['expenses'])

        return data

    def parse_assets_data(self, sheet, assets_data):
        """자산 데이터 파싱 (A~H열)"""
        current_section = None

        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            for col in range(1, min(sheet.max_column + 1, 9)):  # A~H열
                cell_value = sheet.cell(row=row_num, column=col).value
                if cell_value is not None:
                    row_data.append(str(cell_value).strip())
                else:
                    row_data.append('')

            # 빈 행 스킵
            if all(cell == '' for cell in row_data):
                continue

            first_cell = row_data[0]

            # 섹션 구분
            if '현금성 자산' in first_cell:
                current_section = 'cash_assets'
                continue
            elif '저축 + 투자 자산' in first_cell:
                current_section = 'investment_assets'
                continue
            elif '연금 자산' in first_cell:
                current_section = 'pension_assets'
                continue
            elif '지출' in first_cell or '고정지출' in first_cell:
                current_section = None  # 자산 섹션 종료
                break

            # 헤더 행 스킵
            if first_cell in ['구분', '분류'] or (len(row_data) > 1 and row_data[1] == '계좌'):
                continue

            # 자산 데이터 수집
            non_empty_count = len([x for x in row_data if x != ''])
            if current_section and non_empty_count >= 3:
                try:
                    if current_section == 'cash_assets' and first_cell in ['현금', '통장']:
                        balance = self.clean_number(row_data[3])
                        if balance > 0:
                            assets_data[current_section].append({
                                'type': '현금' if first_cell == '현금' else '통장',
                                'account_name': row_data[1] if row_data[1] != '-' else None,
                                'item_name': self.clean_text(row_data[2]),
                                'balance': balance
                            })

                    elif current_section == 'investment_assets' and first_cell in ['저축', '혼합', '주식']:
                        balance = self.clean_number(row_data[3])
                        if balance > 0:
                            assets_data[current_section].append({
                                'category': first_cell,
                                'account_name': row_data[1] if row_data[1] != '-' else None,
                                'item_name': self.clean_text(row_data[2]),
                                'current_value': balance,
                                'deposit_amount': balance
                            })

                    elif current_section == 'pension_assets' and first_cell in ['연금저축', '퇴직연금']:
                        current_value = self.clean_number(row_data[3])
                        deposit_amount = self.clean_number(row_data[4]) if len(row_data) > 4 else current_value
                        if current_value > 0:
                            assets_data[current_section].append({
                                'type': first_cell,
                                'account_name': row_data[1] if row_data[1] != '-' else None,
                                'item_name': self.clean_text(row_data[2]),
                                'current_value': current_value,
                                'contribution': deposit_amount
                            })
                except Exception as e:
                    print(f"      ⚠️ 자산 파싱 오류 (행 {row_num}): {e}")
                    continue

    def parse_expenses_data(self, sheet, expenses_data):
        """지출 데이터 파싱 (I~M열)"""
        current_section = None

        for row_num in range(1, sheet.max_row + 1):
            # I열부터 M열까지 데이터 읽기
            i_cell = sheet.cell(row=row_num, column=9).value   # I열 (분류/항목명)
            j_cell = sheet.cell(row=row_num, column=10).value  # J열 (금액)
            k_cell = sheet.cell(row=row_num, column=11).value  # K열 (결제일)
            l_cell = sheet.cell(row=row_num, column=12).value  # L열 (수단)

            # None 값 처리
            i_val = str(i_cell).strip() if i_cell is not None else ''
            j_val = str(j_cell).strip() if j_cell is not None else ''
            k_val = str(k_cell).strip() if k_cell is not None else ''
            l_val = str(l_cell).strip() if l_cell is not None else ''

            # 섹션 구분자 확인
            if '고정지출(예정)' in i_val or '고정지출(예상)' in i_val or '- 고정지출(예상)' in i_val:
                current_section = 'fixed_expenses'
                continue
            elif '고정지출(선납)' in i_val or '- 고정지출(선납)' in i_val:
                current_section = 'prepaid_expenses'
                continue
            elif i_val in ['분류', '항목', '합계'] or not i_val:
                continue

            # 실제 지출 데이터 파싱
            if current_section and i_val and i_val != 'None':
                try:
                    amount = self.clean_number(j_val)
                    if amount <= 0:
                        continue

                    payment_date = self.clean_payment_date(k_val)
                    payment_method = self.clean_payment_method(l_val)
                    category = self.extract_category(i_val)

                    expense_data = {
                        'category': category,
                        'item_name': i_val,
                        'amount': amount,
                        'payment_date': payment_date,
                        'payment_method': payment_method
                    }

                    expenses_data[current_section].append(expense_data)

                except Exception as e:
                    print(f"      ⚠️ 지출 파싱 오류 (행 {row_num}): {e}")
                    continue

    def clean_number(self, value):
        """숫자 정리 (문자열에서 숫자만 추출)"""
        if value is None or value == '':
            return 0

        numeric_str = re.sub(r'[^\d.]', '', str(value))
        try:
            return int(float(numeric_str)) if numeric_str else 0
        except:
            return 0

    def clean_text(self, text):
        """텍스트 정리"""
        if not text:
            return ''
        cleaned = re.sub(r'\([^)]*\)', '', text)
        return cleaned.strip()[:200]

    def clean_payment_date(self, date_str):
        """결제일 정리 (1-31 숫자만 추출)"""
        if not date_str or date_str == '-' or date_str == 'None':
            return None

        numbers = re.findall(r'\d+', str(date_str))
        if numbers:
            date = int(numbers[0])
            return date if 1 <= date <= 31 else None
        return None

    def clean_payment_method(self, method_str):
        """결제수단 정리"""
        if not method_str or method_str == 'None':
            return '신용'

        method = str(method_str).strip().lower()
        if '신용' in method:
            return '신용'
        elif '체크' in method:
            return '체크'
        elif '현금' in method:
            return '현금'
        else:
            return '신용'

    def extract_category(self, item_name):
        """항목명에서 카테고리 추출"""
        name = item_name.lower()

        if '보험' in name or '라이프' in name or '생명' in name:
            return '보험'
        elif '청약' in name:
            return '저축'
        elif '통신' in name or '인터넷' in name or 'tv' in name or '에르엘' in name or '딜라이브' in name:
            return '통신'
        elif '월세' in name or '리브엠' in name:
            return '주거'
        elif '요금' in name or '전기' in name or '가스' in name or '수도' in name or 'tv수신료' in name:
            return '주거'
        elif '클라우드' in name or '서버' in name or '구글' in name or '클로드' in name:
            return 'IT'
        else:
            return '생활'

    def clear_all_archives(self):
        """모든 아카이브 데이터 정리"""
        print("🧹 기존 아카이브 데이터 정리 중...")

        cursor = self.connection.cursor()
        try:
            # 아카이브 데이터 삭제
            cursor.execute("DELETE FROM assets_archive_data")
            cursor.execute("DELETE FROM fixed_expenses_archive")
            cursor.execute("DELETE FROM prepaid_expenses_archive")
            cursor.execute("DELETE FROM monthly_archives")

            self.connection.commit()
            print("   ✅ 기존 아카이브 데이터 정리 완료")
        except Exception as e:
            print(f"   ❌ 아카이브 정리 오류: {e}")
            self.connection.rollback()
        finally:
            cursor.close()

    def insert_current_month_data(self, current_data):
        """최신 월 데이터를 실시간 테이블에 입력"""
        print(f"💾 {current_data['year']}년 {current_data['month']}월 실시간 테이블 입력 중...")

        cursor = self.connection.cursor()

        try:
            # 기존 실시간 데이터 삭제
            cursor.execute("DELETE FROM cash_assets")
            cursor.execute("DELETE FROM investment_assets")
            cursor.execute("DELETE FROM pension_assets")
            cursor.execute("DELETE FROM fixed_expenses")
            cursor.execute("DELETE FROM prepaid_expenses")

            # 자산 데이터 입력
            assets = current_data['assets']

            # 현금성 자산
            for idx, asset in enumerate(assets['cash_assets']):
                cursor.execute("""
                    INSERT INTO cash_assets (type, account_name, item_name, balance, display_order)
                    VALUES (%s, %s, %s, %s, %s)
                """, (asset['type'], asset['account_name'], asset['item_name'],
                     asset['balance'], idx + 1))

            # 투자 자산
            for idx, asset in enumerate(assets['investment_assets']):
                cursor.execute("""
                    INSERT INTO investment_assets (category, account_name, item_name,
                                                 current_value, deposit_amount, display_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (asset['category'], asset['account_name'], asset['item_name'],
                     asset['current_value'], asset['deposit_amount'], idx + 1))

            # 연금 자산 (account_name 컬럼이 없음)
            for idx, asset in enumerate(assets['pension_assets']):
                cursor.execute("""
                    INSERT INTO pension_assets (type, item_name,
                                              current_value, deposit_amount, display_order)
                    VALUES (%s, %s, %s, %s, %s)
                """, (asset['type'], asset['item_name'],
                     asset['current_value'], asset['contribution'], idx + 1))

            # 지출 데이터 입력
            expenses = current_data['expenses']

            # 고정지출
            for expense in expenses['fixed_expenses']:
                cursor.execute("""
                    INSERT INTO fixed_expenses
                    (category, item_name, amount, payment_date, payment_method, is_active)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    expense['category'],
                    expense['item_name'],
                    expense['amount'],
                    expense['payment_date'],
                    expense['payment_method'],
                    1
                ))

            # 선납지출
            for expense in expenses['prepaid_expenses']:
                cursor.execute("""
                    INSERT INTO prepaid_expenses
                    (item_name, amount, payment_date, payment_method)
                    VALUES (%s, %s, %s, %s)
                """, (
                    expense['item_name'],
                    expense['amount'],
                    expense['payment_date'],
                    expense['payment_method']
                ))

            self.connection.commit()

            # 결과 출력
            print(f"   ✅ 자산: 현금{len(assets['cash_assets'])}개, 투자{len(assets['investment_assets'])}개, 연금{len(assets['pension_assets'])}개")
            print(f"   ✅ 지출: 고정{len(expenses['fixed_expenses'])}개, 선납{len(expenses['prepaid_expenses'])}개")

        except Exception as e:
            print(f"   ❌ 실시간 데이터 입력 오류: {e}")
            self.connection.rollback()
        finally:
            cursor.close()

    def create_archive_snapshots(self, all_data, exclude_current=True):
        """과거 월들을 아카이브 시스템으로 생성"""
        print("🗄️ 월별 아카이브 스냅샷 생성 중...")

        cursor = self.connection.cursor()

        try:
            # 각 월별로 아카이브 생성
            for sheet_name, month_data in all_data.items():
                if exclude_current and sheet_name == self.current_month:
                    continue

                year = month_data['year']
                month = month_data['month']

                print(f"   📅 {year}년 {month}월 아카이브 생성...")

                # 1. monthly_archives 테이블에 월 레코드 생성
                archive_month = f"{year:04d}-{month:02d}-01"
                cursor.execute("""
                    INSERT INTO monthly_archives (archive_month, modification_notes)
                    VALUES (%s, %s)
                """, (archive_month, f"Excel 원본 데이터 - {year}년 {month}월"))

                archive_id = cursor.lastrowid

                # 2. 자산 데이터를 새로운 아카이브 시스템에 저장
                self.save_assets_to_archive(cursor, archive_id, month_data['assets'])

                # 3. 지출 데이터를 기존 아카이브 시스템에 저장
                self.save_expenses_to_archive(cursor, archive_id, month_data['expenses'])

            self.connection.commit()
            archive_count = len(all_data) - (1 if exclude_current else 0)
            print(f"   ✅ 총 {archive_count}개월 아카이브 생성 완료")

        except Exception as e:
            print(f"   ❌ 아카이브 생성 오류: {e}")
            self.connection.rollback()
        finally:
            cursor.close()

    def save_assets_to_archive(self, cursor, archive_id, assets_data):
        """자산 데이터를 새로운 아카이브 시스템에 저장"""

        # 현금성 자산
        for asset in assets_data['cash_assets']:
            asset_data = {
                'id': None,
                'type': asset['type'],
                'account_name': asset['account_name'],
                'item_name': asset['item_name'],
                'balance': asset['balance'],
                'display_order': 0,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'cash_assets', json.dumps(asset_data, ensure_ascii=False)))

        # 투자 자산
        for asset in assets_data['investment_assets']:
            asset_data = {
                'id': None,
                'category': asset['category'],
                'account_name': asset['account_name'],
                'item_name': asset['item_name'],
                'current_value': asset['current_value'],
                'deposit_amount': asset['deposit_amount'],
                'display_order': 0,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'investment_assets', json.dumps(asset_data, ensure_ascii=False)))

        # 연금 자산 (account_name 컬럼이 없음)
        for asset in assets_data['pension_assets']:
            asset_data = {
                'id': None,
                'type': asset['type'],
                'item_name': asset['item_name'],
                'current_value': asset['current_value'],
                'deposit_amount': asset['contribution'],
                'display_order': 0,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'pension_assets', json.dumps(asset_data, ensure_ascii=False)))

    def save_expenses_to_archive(self, cursor, archive_id, expenses_data):
        """지출 데이터를 기존 아카이브 시스템에 저장"""

        # 고정지출
        for expense in expenses_data['fixed_expenses']:
            cursor.execute("""
                INSERT INTO fixed_expenses_archive
                (archive_id, category, item_name, amount, payment_date, payment_method)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                archive_id,
                expense['category'],
                expense['item_name'],
                expense['amount'],
                expense['payment_date'],
                expense['payment_method']
            ))

        # 선납지출
        for expense in expenses_data['prepaid_expenses']:
            cursor.execute("""
                INSERT INTO prepaid_expenses_archive
                (archive_id, item_name, amount, payment_date, payment_method)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                archive_id,
                expense['item_name'],
                expense['amount'],
                expense['payment_date'],
                expense['payment_method']
            ))

    def convert(self):
        """전체 변환 프로세스 실행"""
        print("🚀 Excel → 완전한 아카이브 시스템 변환 시작")

        # 데이터베이스 연결
        if not self.connect_db():
            return False

        try:
            # 기존 아카이브 데이터 정리
            self.clear_all_archives()

            # Excel 데이터 파싱
            all_data = self.parse_excel_data()

            if not all_data:
                print("❌ 파싱된 데이터가 없습니다.")
                return False

            # 최신 월 실시간 테이블에 입력
            if self.current_month in all_data:
                self.insert_current_month_data(all_data[self.current_month])
            else:
                print(f"⚠️ 최신 월 데이터({self.current_month})를 찾을 수 없습니다.")

            # 과거 월들 아카이브 시스템으로 생성
            self.create_archive_snapshots(all_data, exclude_current=True)

            print("🎉 완전한 아카이브 시스템으로 변환 완료!")
            print("\n📋 변환 결과:")
            print(f"   📊 총 {len(all_data)}개월 데이터 처리")
            print(f"   📅 현재 월({self.current_month}): 실시간 테이블")
            archive_count = len(all_data) - (1 if self.current_month in all_data else 0)
            print(f"   🗄️ 아카이브: {archive_count}개월")
            print(f"   💾 자산 아카이브: 새로운 시스템 (assets_archive_data)")
            print(f"   💾 지출 아카이브: 기존 시스템 (*_expenses_archive)")
            print("\n💡 이제 모든 페이지에서 월별 아카이브 조회가 가능합니다!")

            return True

        except Exception as e:
            print(f"❌ 변환 중 오류 발생: {e}")
            return False
        finally:
            if self.connection:
                self.connection.close()

if __name__ == "__main__":
    converter = ExcelDataConverter()
    converter.convert()