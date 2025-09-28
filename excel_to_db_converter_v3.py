#!/usr/bin/env python3
"""
Excel 자산관리시트 데이터를 새로운 아카이브 시스템으로 변환하는 스크립트 v3
- 최신 월: 실시간 테이블 (자산 + 지출 데이터)
- 과거 월들: 새로운 아카이브 시스템 (monthly_archives + assets_archive_data)
- 고정지출/선납지출 데이터 추가 지원
"""

import openpyxl
import mysql.connector
import re
import json
from datetime import datetime
import os

class ExcelToDBConverterV3:
    def __init__(self):
        self.excel_file = '자산관리시트.xlsx'
        self.connection = None
        # 현재 월 설정 (가장 최신 데이터로 사용할 월)
        self.current_month = '202509'  # 필요시 수정

    def connect_db(self):
        """데이터베이스 연결"""
        try:
            self.connection = mysql.connector.connect(
                host='localhost',
                port=3306,
                user='root',
                password='money_dev_2024',
                database='money_management',
                charset='utf8mb4'
            )
            print("✅ 데이터베이스 연결 성공")
            return True
        except Exception as e:
            print(f"❌ 데이터베이스 연결 실패: {e}")
            return False

    def parse_excel_data(self):
        """Excel 파일에서 모든 월 데이터 파싱 (자산 + 지출)"""
        if not os.path.exists(self.excel_file):
            print(f"❌ Excel 파일을 찾을 수 없습니다: {self.excel_file}")
            return {}

        workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
        all_data = {}

        print("📊 Excel 데이터 파싱 중...")

        for sheet_name in workbook.sheetnames:
            if len(sheet_name) == 6 and sheet_name.isdigit():  # YYYYMM 형식
                year = int(sheet_name[:4])
                month = int(sheet_name[4:])
                print(f"   📅 {year}년 {month}월 파싱 중...")

                data = self.parse_sheet_data(workbook[sheet_name])
                all_data[sheet_name] = {
                    'year': year,
                    'month': month,
                    'data': data
                }

        print(f"✅ 총 {len(all_data)}개월 데이터 파싱 완료")
        return all_data

    def parse_sheet_data(self, sheet):
        """개별 시트 데이터 파싱 (자산 + 지출)"""
        data = {
            'cash_assets': [],
            'investment_assets': [],
            'pension_assets': [],
            'fixed_expenses': [],
            'prepaid_expenses': []
        }

        current_section = None

        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            # 더 많은 열 확인 (N열까지)
            for col in range(1, min(sheet.max_column + 1, 15)):
                cell_value = sheet.cell(row=row_num, column=col).value
                if cell_value is not None:
                    row_data.append(str(cell_value).strip())
                else:
                    row_data.append('')

            # 빈 행 스킵
            if all(cell == '' for cell in row_data):
                continue

            first_cell = row_data[0]
            i_cell = row_data[8] if len(row_data) > 8 else ''  # I열 (9번째)

            # 섹션 구분
            if '현금성 자산' in first_cell:
                current_section = 'cash_assets'
                print(f"      🏷️  섹션 변경: {current_section}")
                continue
            elif '저축 + 투자 자산' in first_cell:
                current_section = 'investment_assets'
                print(f"      🏷️  섹션 변경: {current_section}")
                continue
            elif '연금 자산' in first_cell:
                current_section = 'pension_assets'
                print(f"      🏷️  섹션 변경: {current_section}")
                continue
            elif '- 지출' in first_cell:
                current_section = 'daily_expenses'  # 실제 지출 내역 (현재는 스킵)
                print(f"      🏷️  섹션 변경: 일일 지출 (현재 버전에서는 스킵)")
                continue
            elif '고정지출(선납)' in i_cell:
                current_section = 'fixed_expenses'
                print(f"      🏷️  섹션 변경: 고정지출(선납)")
                continue

            # 헤더 행 스킵
            if first_cell in ['구분', '분류'] or (len(row_data) > 1 and row_data[1] == '계좌'):
                print(f"      ⏭️  헤더 행 스킵: {first_cell}")
                continue

            # 데이터 수집
            non_empty_count = len([x for x in row_data if x != ''])
            if current_section and non_empty_count >= 3:
                print(f"      🔍 데이터 처리 시도: 섹션={current_section}, 첫번째={first_cell}, 비어있지않은셀={non_empty_count}")
                try:
                    # 자산 데이터 처리 (기존 로직)
                    if current_section == 'cash_assets' and first_cell in ['현금', '통장']:
                        balance = self.clean_number(row_data[3])
                        if balance > 0:
                            data[current_section].append({
                                'type': '현금' if first_cell == '현금' else '통장',
                                'account_name': row_data[1] if row_data[1] != '-' else None,
                                'item_name': self.clean_text(row_data[2]),
                                'balance': balance
                            })
                            print(f"      ✅ 현금성 자산: {row_data[2]} = {balance:,}원")

                    elif current_section == 'investment_assets' and first_cell in ['저축', '혼합', '주식']:
                        balance = self.clean_number(row_data[3])
                        if balance > 0:
                            data[current_section].append({
                                'category': first_cell,
                                'account_name': row_data[1] if row_data[1] != '-' else None,
                                'item_name': self.clean_text(row_data[2]),
                                'current_value': balance,
                                'deposit_amount': balance
                            })
                            print(f"      ✅ 투자 자산({first_cell}): {row_data[2]} = {balance:,}원")

                    elif current_section == 'pension_assets' and first_cell in ['연금저축', '퇴직연금']:
                        current_value = self.clean_number(row_data[3])
                        deposit_amount = self.clean_number(row_data[4]) if len(row_data) > 4 else current_value
                        if current_value > 0:
                            data[current_section].append({
                                'type': first_cell,
                                'account_name': row_data[1] if row_data[1] != '-' else None,
                                'item_name': self.clean_text(row_data[2]),
                                'current_value': current_value,
                                'contribution': deposit_amount
                            })
                            print(f"      ✅ 연금 자산: {row_data[2]} = {current_value:,}원")

                    # 고정지출 데이터 처리 (신규)
                    elif current_section == 'fixed_expenses' and i_cell and '합계' not in i_cell:
                        # I열(항목명), J열(금액), K열(결제일), L열(수단)
                        item_name = self.clean_text(i_cell)
                        amount = self.clean_number(row_data[9]) if len(row_data) > 9 else 0  # J열
                        payment_date = self.clean_payment_date(row_data[10]) if len(row_data) > 10 else None  # K열
                        payment_method = self.clean_payment_method(row_data[11]) if len(row_data) > 11 else None  # L열

                        if item_name and amount > 0:
                            # 카테고리 추출 (괄호 안의 정보나 키워드로 판단)
                            category = self.extract_category(item_name)

                            data['fixed_expenses'].append({
                                'category': category,
                                'item_name': item_name,
                                'amount': amount,
                                'payment_date': payment_date,
                                'payment_method': payment_method
                            })
                            print(f"      ✅ 고정지출: {item_name} = {amount:,}원 ({payment_date}일, {payment_method})")

                except Exception as e:
                    print(f"      ⚠️  행 {row_num} 파싱 오류: {e}")
                    continue

        return data

    def clean_number(self, value):
        """숫자 정리 (문자열에서 숫자만 추출)"""
        if value is None or value == '':
            return 0

        # 문자열에서 숫자와 소수점만 추출
        numeric_str = re.sub(r'[^\d.]', '', str(value))
        try:
            return int(float(numeric_str)) if numeric_str else 0
        except:
            return 0

    def clean_text(self, text):
        """텍스트 정리 (특수문자 제거 등)"""
        if not text:
            return ''

        # 괄호 안 내용은 유지 (보험 정보 등이 중요함)
        cleaned = text.strip()
        return cleaned[:200]  # 최대 200자

    def clean_payment_date(self, date_str):
        """결제일 정리 (1-31 숫자만 추출)"""
        if not date_str or date_str == '-':
            return None

        # 숫자만 추출
        numbers = re.findall(r'\d+', str(date_str))
        if numbers:
            date = int(numbers[0])
            return date if 1 <= date <= 31 else None
        return None

    def clean_payment_method(self, method_str):
        """결제수단 정리 (신용/체크/현금)"""
        if not method_str:
            return None

        method = str(method_str).strip().lower()
        if '신용' in method:
            return '신용'
        elif '체크' in method:
            return '체크'
        elif '현금' in method:
            return '현금'
        else:
            return '신용'  # 기본값

    def extract_category(self, item_name):
        """항목명에서 카테고리 추출"""
        name = item_name.lower()

        if '보험' in name or '라이프' in name or '생명' in name:
            return '보험'
        elif '청약' in name:
            return '저축'
        elif '통신' in name or '인터넷' in name or 'tv' in name:
            return '통신'
        elif '월세' in name or '리브엠' in name:
            return '주거'
        elif '요금' in name:
            return '공공요금'
        else:
            return '기타'

    def insert_current_month_data(self, current_data):
        """최신 월 데이터를 실시간 테이블에 입력 (자산 + 지출)"""
        print(f"💾 {current_data['year']}년 {current_data['month']}월 실시간 테이블 입력 중...")

        cursor = self.connection.cursor()

        try:
            # 기존 데이터 삭제 (깨끗하게 시작)
            cursor.execute("DELETE FROM cash_assets")
            cursor.execute("DELETE FROM investment_assets")
            cursor.execute("DELETE FROM pension_assets")
            cursor.execute("DELETE FROM fixed_expenses")
            cursor.execute("DELETE FROM prepaid_expenses")

            # 현금성 자산 입력
            for idx, asset in enumerate(current_data['data']['cash_assets']):
                cursor.execute("""
                    INSERT INTO cash_assets (type, account_name, item_name, balance, display_order)
                    VALUES (%s, %s, %s, %s, %s)
                """, (asset['type'], asset['account_name'], asset['item_name'],
                     asset['balance'], idx + 1))

            # 투자 자산 입력
            for idx, asset in enumerate(current_data['data']['investment_assets']):
                cursor.execute("""
                    INSERT INTO investment_assets (category, account_name, item_name,
                                                 current_value, deposit_amount, display_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (asset['category'], asset['account_name'], asset['item_name'],
                     asset['current_value'], asset['deposit_amount'], idx + 1))

            # 연금 자산 입력
            for idx, asset in enumerate(current_data['data']['pension_assets']):
                cursor.execute("""
                    INSERT INTO pension_assets (type, account_name, item_name,
                                              current_value, deposit_amount, display_order)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (asset['type'], asset['account_name'], asset['item_name'],
                     asset['current_value'], asset['contribution'], idx + 1))

            # 고정지출 입력 (신규)
            for expense in current_data['data']['fixed_expenses']:
                cursor.execute("""
                    INSERT INTO fixed_expenses (category, item_name, amount, payment_date, payment_method)
                    VALUES (%s, %s, %s, %s, %s)
                """, (expense['category'], expense['item_name'], expense['amount'],
                     expense['payment_date'], expense['payment_method']))

            self.connection.commit()

            # 결과 출력
            cash_count = len(current_data['data']['cash_assets'])
            investment_count = len(current_data['data']['investment_assets'])
            pension_count = len(current_data['data']['pension_assets'])
            fixed_expense_count = len(current_data['data']['fixed_expenses'])

            print(f"   ✅ 현금성 자산: {cash_count}개")
            print(f"   ✅ 투자 자산: {investment_count}개")
            print(f"   ✅ 연금 자산: {pension_count}개")
            print(f"   ✅ 고정지출: {fixed_expense_count}개")

        except Exception as e:
            print(f"   ❌ 실시간 데이터 입력 오류: {e}")
            self.connection.rollback()
        finally:
            cursor.close()

    def create_archive_snapshots(self, all_data, exclude_current=True):
        """과거 월들을 새로운 아카이브 시스템으로 생성"""
        print("🗄️  새로운 아카이브 스냅샷 생성 중...")

        cursor = self.connection.cursor()

        try:
            # 기존 아카이브 데이터 삭제
            cursor.execute("DELETE FROM assets_archive_data")
            cursor.execute("DELETE FROM monthly_archives")

            # 각 월별로 아카이브 생성
            for sheet_name, month_data in all_data.items():
                if exclude_current and sheet_name == self.current_month:
                    continue

                year = month_data['year']
                month = month_data['month']
                data = month_data['data']

                print(f"   📅 {year}년 {month}월 아카이브 생성...")

                # 1. monthly_archives 테이블에 월 레코드 생성
                archive_month = f"{year:04d}-{month:02d}-01"
                cursor.execute("""
                    INSERT INTO monthly_archives (archive_month, modification_notes)
                    VALUES (%s, %s)
                """, (archive_month, f"Excel 데이터 변환 - {year}년 {month}월"))

                # 생성된 archive_id 가져오기
                archive_id = cursor.lastrowid

                # 2. assets_archive_data 테이블에 자산 + 지출 데이터 저장
                self.save_data_to_archive(cursor, archive_id, data)

            self.connection.commit()
            archive_count = len(all_data) - (1 if exclude_current else 0)
            print(f"   ✅ 총 {archive_count}개월 아카이브 생성 완료")

        except Exception as e:
            print(f"   ❌ 아카이브 생성 오류: {e}")
            self.connection.rollback()
        finally:
            cursor.close()

    def save_data_to_archive(self, cursor, archive_id, data):
        """새로운 아카이브 시스템에 자산 + 지출 데이터 저장"""

        # 현금성 자산 아카이브
        for asset in data['cash_assets']:
            asset_data = {
                'id': None,
                'type': asset['type'],
                'account_name': asset['account_name'],
                'item_name': asset['item_name'],
                'balance': asset['balance'],
                'display_order': 0,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'cash_assets', json.dumps(asset_data, ensure_ascii=False)))

        # 투자 자산 아카이브
        for asset in data['investment_assets']:
            asset_data = {
                'id': None,
                'category': asset['category'],
                'account_name': asset['account_name'],
                'item_name': asset['item_name'],
                'current_value': asset['current_value'],
                'deposit_amount': asset['deposit_amount'],
                'display_order': 0,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'investment_assets', json.dumps(asset_data, ensure_ascii=False)))

        # 연금 자산 아카이브
        for asset in data['pension_assets']:
            asset_data = {
                'id': None,
                'type': asset['type'],
                'account_name': asset['account_name'],
                'item_name': asset['item_name'],
                'current_value': asset['current_value'],
                'deposit_amount': asset['contribution'],
                'display_order': 0,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'pension_assets', json.dumps(asset_data, ensure_ascii=False)))

        # 고정지출 아카이브 (신규)
        for expense in data['fixed_expenses']:
            expense_data = {
                'id': None,
                'category': expense['category'],
                'item_name': expense['item_name'],
                'amount': expense['amount'],
                'payment_date': expense['payment_date'],
                'payment_method': expense['payment_method'],
                'is_active': True,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'deleted_at': None
            }

            cursor.execute("""
                INSERT INTO assets_archive_data (archive_id, asset_table, asset_data)
                VALUES (%s, %s, %s)
            """, (archive_id, 'fixed_expenses', json.dumps(expense_data, ensure_ascii=False)))

    def convert(self):
        """전체 변환 프로세스 실행"""
        print("🚀 Excel → 새로운 아카이브 시스템 변환 시작 (자산 + 지출 데이터)")

        # 데이터베이스 연결
        if not self.connect_db():
            return False

        try:
            # Excel 데이터 파싱
            all_data = self.parse_excel_data()

            if not all_data:
                print("❌ 파싱된 데이터가 없습니다.")
                return False

            # 최신 월 실시간 테이블에 입력
            if self.current_month in all_data:
                self.insert_current_month_data(all_data[self.current_month])
            else:
                print(f"⚠️ 최신 월 데이터({self.current_month})를 찾을 수 없습니다.")

            # 과거 월들 새로운 아카이브 시스템으로 생성
            self.create_archive_snapshots(all_data, exclude_current=True)

            print("🎉 새로운 아카이브 시스템으로 변환 완료!")
            print("\n📋 변환 결과:")
            print(f"   📊 총 {len(all_data)}개월 데이터 처리")
            print(f"   📅 현재 월({self.current_month}): 실시간 테이블 (자산 + 지출)")
            archive_count = len(all_data) - (1 if self.current_month in all_data else 0)
            print(f"   🗄️ 아카이브: {archive_count}개월")
            print("\n💡 이제 자산현황 페이지와 지출현황 페이지에서 월별 아카이브 조회가 가능합니다!")

            return True

        except Exception as e:
            print(f"❌ 변환 중 오류 발생: {e}")
            return False
        finally:
            if self.connection:
                self.connection.close()

if __name__ == "__main__":
    converter = ExcelToDBConverterV3()
    converter.convert()